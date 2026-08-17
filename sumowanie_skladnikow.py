import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
import threading
import os
import tempfile
import subprocess
import time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from tkinterdnd2 import TkinterDnD, DND_FILES


class ExcelSumApp:

    def __init__(self, root):

        self.root = root

        self.root.title(
            "Sumowanie składników Excel"
        )

        self.root.geometry(
            "1300x850"
        )

        self.root.minsize(
            1000,
            700
        )

        self.files = []
        self.rows = []

        self.create_styles()
        self.create_interface()

    # ==========================================================
    # STYLE
    # ==========================================================

    def create_styles(self):

        style = ttk.Style()

        try:
            style.theme_use("vista")
        except tk.TclError:
            pass

        style.configure(
            "Title.TLabel",
            font=("Segoe UI", 18, "bold")
        )

        style.configure(
            "Big.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(15, 9)
        )

        style.configure(
            "Treeview",
            font=("Segoe UI", 10),
            rowheight=28
        )

        style.configure(
            "Treeview.Heading",
            font=("Segoe UI", 10, "bold")
        )

    # ==========================================================
    # INTERFEJS
    # ==========================================================

    def create_interface(self):

        main = ttk.Frame(
            self.root,
            padding=18
        )

        main.pack(
            fill="both",
            expand=True
        )

        # ======================================================
        # NAGŁÓWEK
        # ======================================================

        header = ttk.Frame(main)

        header.pack(
            fill="x",
            pady=(0, 12)
        )

        ttk.Label(
            header,
            text="Sumowanie składników Excel",
            style="Title.TLabel"
        ).pack(
            side="left"
        )

        ttk.Button(
            header,
            text="Wybierz pliki Excel",
            command=self.choose_files,
            style="Big.TButton"
        ).pack(
            side="right"
        )

        # ======================================================
        # DRAG & DROP
        # ======================================================

        self.drop_area = tk.Frame(
            main,
            bd=2,
            relief="groove",
            height=100
        )

        self.drop_area.pack(
            fill="x",
            pady=(0, 12)
        )

        self.drop_area.pack_propagate(
            False
        )

        self.drop_label = tk.Label(
            self.drop_area,
            text=(
                "PRZECIĄGNIJ TUTAJ PLIKI EXCEL\n\n"
                "Możesz przeciągnąć kilka plików .xlsx jednocześnie"
            ),
            font=("Segoe UI", 11, "bold"),
            justify="center"
        )

        self.drop_label.pack(
            fill="both",
            expand=True
        )

        self.drop_area.drop_target_register(
            DND_FILES
        )

        self.drop_area.dnd_bind(
            "<<Drop>>",
            self.drop_files
        )

        self.drop_label.drop_target_register(
            DND_FILES
        )

        self.drop_label.dnd_bind(
            "<<Drop>>",
            self.drop_files
        )

        # ======================================================
        # LISTA PLIKÓW
        # ======================================================

        ttk.Label(
            main,
            text="Wczytane pliki:"
        ).pack(
            anchor="w"
        )

        files_frame = ttk.Frame(
            main
        )

        files_frame.pack(
            fill="x",
            pady=(4, 10)
        )

        self.file_list = tk.Listbox(
            files_frame,
            height=5,
            font=("Segoe UI", 10)
        )

        self.file_list.pack(
            side="left",
            fill="both",
            expand=True
        )

        files_scroll = ttk.Scrollbar(
            files_frame,
            orient="vertical",
            command=self.file_list.yview
        )

        files_scroll.pack(
            side="right",
            fill="y"
        )

        self.file_list.configure(
            yscrollcommand=files_scroll.set
        )

        # ======================================================
        # PRZYCISKI PLIKÓW
        # ======================================================

        file_buttons = ttk.Frame(
            main
        )

        file_buttons.pack(
            fill="x",
            pady=(0, 10)
        )

        ttk.Button(
            file_buttons,
            text="Usuń zaznaczony plik",
            command=self.remove_selected_file
        ).pack(
            side="left"
        )

        ttk.Button(
            file_buttons,
            text="Wyczyść listę",
            command=self.clear
        ).pack(
            side="left",
            padx=(8, 0)
        )

        ttk.Button(
            file_buttons,
            text="Przetwórz pliki",
            command=self.start_processing,
            style="Big.TButton"
        ).pack(
            side="right"
        )

        # ======================================================
        # STATUS
        # ======================================================

        self.status = tk.StringVar(
            value="Dodaj pliki Excel."
        )

        ttk.Label(
            main,
            textvariable=self.status
        ).pack(
            fill="x",
            pady=(0, 8)
        )

        # ======================================================
        # TABELA
        # ======================================================

        table_frame = ttk.Frame(
            main
        )

        table_frame.pack(
            fill="both",
            expand=True
        )

        columns = (
            "number",
            "alternative",
            "description",
            "quantity",
            "unit"
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings"
        )

        self.tree.heading(
            "number",
            text="Numer składnika"
        )

        self.tree.heading(
            "alternative",
            text="Altern."
        )

        self.tree.heading(
            "description",
            text="Opis składnika"
        )

        self.tree.heading(
            "quantity",
            text="Ilość"
        )

        self.tree.heading(
            "unit",
            text="Jednostka"
        )

        self.tree.column(
            "number",
            width=160,
            anchor="center"
        )

        # Wąska kolumna X/Y

        self.tree.column(
            "alternative",
            width=75,
            minwidth=60,
            anchor="center",
            stretch=False
        )

        self.tree.column(
            "description",
            width=620,
            anchor="w"
        )

        self.tree.column(
            "quantity",
            width=150,
            anchor="e"
        )

        self.tree.column(
            "unit",
            width=120,
            anchor="center"
        )

        scroll_y = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.tree.yview
        )

        scroll_x = ttk.Scrollbar(
            table_frame,
            orient="horizontal",
            command=self.tree.xview
        )

        self.tree.configure(
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        self.tree.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        scroll_y.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        scroll_x.grid(
            row=1,
            column=0,
            sticky="ew"
        )

        table_frame.rowconfigure(
            0,
            weight=1
        )

        table_frame.columnconfigure(
            0,
            weight=1
        )

        # ======================================================
        # DOLNY PASEK
        # ======================================================

        bottom = ttk.Frame(
            main
        )

        bottom.pack(
            fill="x",
            pady=(12, 0)
        )

        bottom.columnconfigure(
            0,
            weight=1
        )

        self.count = tk.StringVar(
            value="Pliki: 0 | Wiersze: 0 | Pozycje: 0"
        )

        ttk.Label(
            bottom,
            textvariable=self.count
        ).pack(
            side="left"
        )

        # ------------------------------------------------------
        # SZYBKI DRUK
        # ------------------------------------------------------

        self.print_button = ttk.Button(
            bottom,
            text="Szybki druk",
            command=self.quick_print,
            style="Big.TButton",
            state="disabled"
        )

        self.print_button.pack(
            side="right",
            padx=(8, 0)
        )

        # ------------------------------------------------------
        # ZAPIS
        # ------------------------------------------------------

        self.save_button = ttk.Button(
            bottom,
            text="Zapisz jako Excel",
            command=self.save_file,
            style="Big.TButton",
            state="disabled"
        )

        self.save_button.pack(
            side="right"
        )

    # ==========================================================
    # DRAG & DROP
    # ==========================================================

    def drop_files(
        self,
        event
    ):

        try:

            paths = self.root.tk.splitlist(
                event.data
            )

            added = 0

            for path in paths:

                path = path.strip("{}")

                if not path.lower().endswith(
                    ".xlsx"
                ):
                    continue

                if path in self.files:
                    continue

                self.files.append(
                    path
                )

                self.file_list.insert(
                    tk.END,
                    Path(path).name
                )

                added += 1

            if added == 0:

                messagebox.showwarning(
                    "Brak plików",
                    "Nie znaleziono nowych plików .xlsx."
                )

                return

            self.status.set(
                f"Dodano {added} plików."
            )

        except Exception as exc:

            messagebox.showerror(
                "Błąd",
                str(exc)
            )

    # ==========================================================
    # WYBÓR PLIKÓW
    # ==========================================================

    def choose_files(self):

        paths = filedialog.askopenfilenames(
            title="Wybierz pliki Excel",
            filetypes=[
                (
                    "Pliki Excel",
                    "*.xlsx"
                ),
                (
                    "Wszystkie pliki",
                    "*.*"
                )
            ]
        )

        if not paths:
            return

        added = 0

        for path in paths:

            if path in self.files:
                continue

            self.files.append(
                path
            )

            self.file_list.insert(
                tk.END,
                Path(path).name
            )

            added += 1

        self.status.set(
            f"Dodano {added} plików."
        )

    # ==========================================================
    # USUWANIE PLIKU
    # ==========================================================

    def remove_selected_file(self):

        selected = self.file_list.curselection()

        if not selected:
            return

        index = selected[0]

        self.file_list.delete(
            index
        )

        del self.files[index]

        self.status.set(
            f"Plików na liście: {len(self.files)}"
        )

    # ==========================================================
    # CZYSZCZENIE
    # ==========================================================

    def clear(self):

        self.files = []
        self.rows = []

        self.file_list.delete(
            0,
            tk.END
        )

        for item in self.tree.get_children():

            self.tree.delete(
                item
            )

        self.count.set(
            "Pliki: 0 | Wiersze: 0 | Pozycje: 0"
        )

        self.status.set(
            "Lista wyczyszczona."
        )

        self.save_button.config(
            state="disabled"
        )

        self.print_button.config(
            state="disabled"
        )

    # ==========================================================
    # START
    # ==========================================================

    def start_processing(self):

        if not self.files:

            messagebox.showwarning(
                "Brak plików",
                "Dodaj przynajmniej jeden plik Excel."
            )

            return

        self.status.set(
            f"Przetwarzanie {len(self.files)} plików..."
        )

        self.save_button.config(
            state="disabled"
        )

        self.print_button.config(
            state="disabled"
        )

        threading.Thread(
            target=self.process_files,
            daemon=True
        ).start()

    # ==========================================================
    # LICZBA
    # ==========================================================

    @staticmethod
    def to_decimal(
        value
    ):

        if value is None:
            raise InvalidOperation

        if isinstance(
            value,
            bool
        ):
            raise InvalidOperation

        if isinstance(
            value,
            (
                int,
                float,
                Decimal
            )
        ):

            return Decimal(
                str(value)
            )

        text = str(
            value
        ).strip()

        text = text.replace(
            "\u00a0",
            ""
        )

        text = text.replace(
            " ",
            ""
        )

        text = text.replace(
            ",",
            "."
        )

        return Decimal(
            text
        )

    # ==========================================================
    # NAGŁÓWKI
    # ==========================================================

    @staticmethod
    def is_header(
        value,
        column_type
    ):

        if value is None:
            return False

        text = (
            str(value)
            .strip()
            .lower()
            .replace(
                "\n",
                " "
            )
        )

        if column_type == "number":

            return (
                "numer składnika" in text
                or "numer skladnika" in text
                or "nr składnika" in text
                or "nr skladnika" in text
            )

        if column_type == "alternative":

            return (
                "alternatywa" in text
                or "alternatywa surowca" in text
            )

        if column_type == "description":

            return (
                "opis składnika" in text
                or "opis skladnika" in text
            )

        if column_type == "quantity":

            return (
                "ilość" in text
                or "ilosc" in text
            )

        if column_type == "unit":

            return (
                "jednostka" in text
                or "jm skł" in text
                or "jm skl" in text
            )

        return False

    # ==========================================================
    # START WIERSZA
    # ==========================================================

    def get_start_row(
        self,
        worksheet
    ):

        headers = [
            worksheet.cell(
                1,
                column
            ).value

            for column in (
                2,
                3,
                4,
                5,
                6
            )
        ]

        matches = [

            self.is_header(
                headers[0],
                "number"
            ),

            self.is_header(
                headers[1],
                "alternative"
            ),

            self.is_header(
                headers[2],
                "description"
            ),

            self.is_header(
                headers[3],
                "quantity"
            ),

            self.is_header(
                headers[4],
                "unit"
            )
        ]

        if matches[3]:

            return 2

        if sum(matches) >= 2:

            return 2

        return 1

    # ==========================================================
    # PRZETWARZANIE
    # ==========================================================

    def process_files(
        self
    ):

        grouped = defaultdict(
            Decimal
        )

        errors = []

        total_rows = 0

        for file_path in self.files:

            try:

                workbook = load_workbook(
                    file_path,
                    read_only=True,
                    data_only=True
                )

                worksheet = workbook.active

                start_row = self.get_start_row(
                    worksheet
                )

                for row_number, row in enumerate(

                    worksheet.iter_rows(
                        min_row=start_row,
                        min_col=2,
                        max_col=6,
                        values_only=True
                    ),

                    start=start_row
                ):

                    number = row[0]
                    alternative = row[1]
                    description = row[2]
                    quantity = row[3]
                    unit = row[4]

                    if all(

                        value is None
                        or str(value).strip() == ""

                        for value in (
                            number,
                            alternative,
                            description,
                            quantity,
                            unit
                        )
                    ):

                        continue

                    total_rows += 1

                    if (
                        number is None
                        or str(number).strip() == ""
                    ):

                        errors.append(
                            f"{Path(file_path).name}, "
                            f"wiersz {row_number}: "
                            f"brak Numeru składnika."
                        )

                        continue

                    try:

                        qty = self.to_decimal(
                            quantity
                        )

                    except (
                        InvalidOperation,
                        ValueError
                    ):

                        errors.append(
                            f"{Path(file_path).name}, "
                            f"wiersz {row_number}: "
                            f"'{quantity}' w kolumnie E "
                            f"nie jest liczbą."
                        )

                        continue

                    number = str(
                        number
                    ).strip()

                    alternative = (
                        ""
                        if alternative is None
                        else str(
                            alternative
                        ).strip()
                    )

                    description = (
                        ""
                        if description is None
                        else str(
                            description
                        ).strip()
                    )

                    unit = (
                        ""
                        if unit is None
                        else str(
                            unit
                        ).strip()
                    )

                    # ==================================================
                    # SUMOWANIE:
                    #
                    # NUMER + ALTERNATYWA + OPIS + JEDNOSTKA
                    #
                    # X + X
                    # Y + Y
                    # ==================================================

                    key = (
                        number,
                        alternative,
                        description,
                        unit
                    )

                    grouped[key] += qty

                workbook.close()

            except Exception as exc:

                errors.append(
                    f"{Path(file_path).name}: "
                    f"nie można odczytać pliku — "
                    f"{exc}"
                )

        result = []

        for key, total in grouped.items():

            result.append(
                (
                    key[0],
                    key[1],
                    key[2],
                    total,
                    key[3]
                )
            )

        result.sort(
            key=lambda row: (
                row[0],
                row[1]
            )
        )

        self.root.after(
            0,
            self.show_result,
            result,
            errors,
            total_rows
        )

    # ==========================================================
    # FORMAT LICZBY
    # ==========================================================

    @staticmethod
    def format_number(
        value
    ):

        if value == value.to_integral():

            return str(
                int(value)
            )

        text = format(
            value.normalize(),
            "f"
        )

        if "." in text:

            text = (
                text
                .rstrip("0")
                .rstrip(".")
            )

        return text

    # ==========================================================
    # WYNIK
    # ==========================================================

    def show_result(
        self,
        result,
        errors,
        total_rows
    ):

        self.rows = result

        for item in self.tree.get_children():

            self.tree.delete(
                item
            )

        for (
            number,
            alternative,
            description,
            quantity,
            unit
        ) in result:

            self.tree.insert(
                "",
                "end",
                values=(
                    number,
                    alternative,
                    description,
                    self.format_number(
                        quantity
                    ),
                    unit
                )
            )

        self.count.set(
            f"Pliki: {len(self.files)} | "
            f"Wiersze: {total_rows} | "
            f"Pozycje po zsumowaniu: {len(result)}"
        )

        if result:

            self.save_button.config(
                state="normal"
            )

            self.print_button.config(
                state="normal"
            )

        if errors:

            self.status.set(
                f"Gotowe. Znaleziono "
                f"{len(errors)} problemów."
            )

            preview = "\n".join(
                errors[:20]
            )

            if len(errors) > 20:

                preview += (
                    f"\n... oraz "
                    f"{len(errors) - 20} kolejnych."
                )

            messagebox.showwarning(
                "Problemy w danych",
                "Pliki zostały przetworzone, "
                "ale znaleziono problemy:\n\n"
                + preview
            )

        else:

            self.status.set(
                "Gotowe. Wszystkie pliki "
                "przetworzono poprawnie."
            )

    # ==========================================================
    # FORMATOWANIE EXCELA
    # ==========================================================

    def format_worksheet(
        self,
        worksheet
    ):

        # ------------------------------------------------------
        # CZCIONKI
        # ------------------------------------------------------

        header_font = Font(
            bold=True
        )

        # ------------------------------------------------------
        # OBRAMOWANIE
        # ------------------------------------------------------

        thin = Side(
            style="thin"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        # ------------------------------------------------------
        # KOMÓRKI
        # ------------------------------------------------------

        for row in worksheet.iter_rows():

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    vertical="center"
                )

        # ------------------------------------------------------
        # NAGŁÓWEK
        # ------------------------------------------------------

        for cell in worksheet[1]:

            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # ------------------------------------------------------
        # SZEROKOŚCI
        # ------------------------------------------------------

        worksheet.column_dimensions[
            "A"
        ].width = 18

        worksheet.column_dimensions[
            "B"
        ].width = 8

        worksheet.column_dimensions[
            "C"
        ].width = 55

        worksheet.column_dimensions[
            "D"
        ].width = 14

        worksheet.column_dimensions[
            "E"
        ].width = 12

        # ------------------------------------------------------
        # ZAMROŻENIE NAGŁÓWKA
        # ------------------------------------------------------

        worksheet.freeze_panes = "A2"

        # ------------------------------------------------------
        # FILTR
        # ------------------------------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        # ------------------------------------------------------
        # USTAWIENIA DRUKOWANIA
        # ------------------------------------------------------

        worksheet.page_setup.orientation = (
            "landscape"
        )

        worksheet.page_setup.paperSize = (
            worksheet.PAPERSIZE_A4
        )

        # Jedna strona szerokości

        worksheet.page_setup.fitToWidth = 1

        # Wysokość może mieć wiele stron

        worksheet.page_setup.fitToHeight = 0

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        # Powtarzaj nagłówek na każdej stronie

        worksheet.print_title_rows = "1:1"

        # Marginesy

        worksheet.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.5,
            bottom=0.5,
            header=0.2,
            footer=0.2
        )

        # Wyśrodkowanie poziome

        worksheet.print_options.horizontalCentered = True

        # Obszar wydruku

        worksheet.print_area = worksheet.dimensions

    # ==========================================================
    # UTWORZENIE EXCELA
    # ==========================================================

    def create_output_workbook(
        self
    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = (
            "Zsumowane składniki"
        )

        worksheet.append(
            [
                "Numer składnika",
                "Alternatywa surowca",
                "Opis składnika",
                "Ilość",
                "Jednostka"
            ]
        )

        for (
            number,
            alternative,
            description,
            quantity,
            unit
        ) in self.rows:

            worksheet.append(
                [
                    number,
                    alternative,
                    description,
                    float(quantity),
                    unit
                ]
            )

        self.format_worksheet(
            worksheet
        )

        return workbook

    # ==========================================================
    # ZAPIS DO EXCELA
    # ==========================================================

    def save_file(
        self
    ):

        if not self.rows:

            messagebox.showwarning(
                "Brak danych",
                "Nie ma danych do zapisania."
            )

            return

        path = filedialog.asksaveasfilename(

            title="Zapisz wynik",

            defaultextension=".xlsx",

            initialfile=(
                "zsumowane_skladniki.xlsx"
            ),

            filetypes=[
                (
                    "Plik Excel",
                    "*.xlsx"
                )
            ]
        )

        if not path:
            return

        try:

            workbook = self.create_output_workbook()

            workbook.save(
                path
            )

            self.status.set(
                "Zapisano wynik."
            )

            messagebox.showinfo(
                "Zapisano",
                "Gotowy plik zapisano jako:\n\n"
                + path
            )

        except PermissionError:

            messagebox.showerror(
                "Brak dostępu",
                "Nie można zapisać pliku. "
                "Sprawdź, czy plik nie jest "
                "otwarty w Excelu."
            )

        except Exception as exc:

            messagebox.showerror(
                "Błąd zapisu",
                str(exc)
            )

    # ==========================================================
    # SZYBKI DRUK
    # ==========================================================

    def quick_print(
        self
    ):

        if not self.rows:

            messagebox.showwarning(
                "Brak danych",
                "Najpierw przetwórz pliki Excel."
            )

            return

        self.status.set(
            "Przygotowywanie wydruku..."
        )

        self.print_button.config(
            state="disabled"
        )

        threading.Thread(
            target=self.print_worker,
            daemon=True
        ).start()

    # ==========================================================
    # DRUK - WĄTEK
    # ==========================================================

    def print_worker(
        self
    ):

        temp_path = None

        try:

            # --------------------------------------------------
            # TEMP
            # --------------------------------------------------

            temp_dir = tempfile.gettempdir()

            temp_path = os.path.join(
                temp_dir,
                "sumowanie_skladnikow_print.xlsx"
            )

            # Usuń poprzednią wersję

            if os.path.exists(
                temp_path
            ):

                try:

                    os.remove(
                        temp_path
                    )

                except Exception:
                    pass

            # --------------------------------------------------
            # UTWÓRZ EXCEL
            # --------------------------------------------------

            workbook = self.create_output_workbook()

            workbook.save(
                temp_path
            )

            workbook.close()

            # --------------------------------------------------
            # DRUKOWANIE PRZEZ WINDOWS
            # --------------------------------------------------

            if not os.path.exists(
                temp_path
            ):

                raise FileNotFoundError(
                    "Nie udało się utworzyć "
                    "tymczasowego pliku Excel."
                )

            # Windows ShellExecute "print"
            #
            # Używa domyślnej aplikacji dla .xlsx
            # i domyślnej drukarki.

            try:

                os.startfile(
                    temp_path,
                    "print"
                )

            except AttributeError:

                # Awaryjnie przez cmd

                subprocess.Popen(
                    [
                        "cmd",
                        "/c",
                        "start",
                        "",
                        "/wait",
                        temp_path
                    ],
                    shell=True
                )

            # --------------------------------------------------
            # CZEKAMY CHWILĘ NA PRZEKAZANIE PLIKU
            # --------------------------------------------------

            time.sleep(
                5
            )

            self.root.after(
                0,
                self.print_finished
            )

            # --------------------------------------------------
            # USUNIĘCIE TEMP
            # --------------------------------------------------

            def remove_temp():

                try:

                    if os.path.exists(
                        temp_path
                    ):

                        os.remove(
                            temp_path
                        )

                except Exception:

                    pass

            threading.Thread(
                target=remove_temp,
                daemon=True
            ).start()

        except Exception as exc:

            self.root.after(
                0,
                self.print_error,
                str(exc)
            )

    # ==========================================================
    # DRUK ZAKOŃCZONY
    # ==========================================================

    def print_finished(
        self
    ):

        self.print_button.config(
            state="normal"
        )

        self.status.set(
            "Wysłano dokument do domyślnej drukarki."
        )

        messagebox.showinfo(
            "Drukowanie",
            "Dokument został wysłany "
            "do domyślnej drukarki Windows."
        )

    # ==========================================================
    # BŁĄD DRUKOWANIA
    # ==========================================================

    def print_error(
        self,
        error
    ):

        self.print_button.config(
            state="normal"
        )

        self.status.set(
            "Błąd drukowania."
        )

        messagebox.showerror(
            "Błąd drukowania",
            "Nie udało się wysłać dokumentu "
            "do drukarki.\n\n"
            + error
        )


# ==============================================================
# START PROGRAMU
# ==============================================================

def main():

    root = TkinterDnD.Tk()

    ExcelSumApp(
        root
    )

    root.mainloop()


if __name__ == "__main__":

    main()